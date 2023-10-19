import argparse
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import datetime

from SERVER.server import BaubuddyAPIHandler

# Define color codes
COLOR_CODES = {
    "green": "007500",
    "orange": "FFA500",
    "red": "b30000"
}

class VehicleDataProcessor(BaubuddyAPIHandler):
    
    def __init__(self):
        super().__init__()
        self.access_token = self.get_access_token()

    def process_response(self, response, keys, colored):
        df = pd.read_csv("vehicles.csv",skiprows=2)
        response = df
        df = pd.DataFrame(response)

        # Filter and sort as required
        df = df[keys + ["gruppe", "rnr"]].sort_values(by="gruppe")

        # Generate Excel
        wb = Workbook()
        ws = wb.active

        # Write header
        header = keys + ["gruppe", "rnr"]
        for col_num, header_text in enumerate(header, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header_text

        # Apply color coding
        if colored:
            for row_num, row in enumerate(df.iterrows(), start=2):
                hu = row[1]["hu"]
                fill_color = None
                if hu <= 90:
                    fill_color = COLOR_CODES["green"]
                elif hu <= 365:
                    fill_color = COLOR_CODES["orange"]
                else:
                    fill_color = COLOR_CODES["red"]
                for col_num, col_value in enumerate(row[1], start=1):
                    col_letter = get_column_letter(col_num)
                    cell = ws[f"{col_letter}{row_num}"]
                    cell.value = col_value
                    if col_num in range(3, len(keys) + 3):  # Apply color only to keys columns
                        cell.fill = PatternFill(start_color=f"{fill_color}FF", end_color=f"{fill_color}FF", fill_type="solid")

        # Save Excel file
        current_date_iso_formatted = datetime.datetime.now().isoformat()[:-7].replace("T", "_").replace(":", "-")
        file_name = f"vehicles_{current_date_iso_formatted}.xlsx"
        wb.save(file_name)
        return file_name

    def process_csv(self, csv_file_path, keys, colored):
        with open(csv_file_path, "rb") as file:
            files = {"file": file}
            headers = {"Authorization": f"Bearer {self.access_token}"}
            response = requests.post(self.API_BASE_URL, files=files, headers=headers).json()

        excel_file = self.process_response(response, keys, colored)
        print(f"Excel file '{excel_file}' generated successfully.")
